#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF-Lesezeichen -> Inhaltsverzeichnis (PDF + Excel) + Master-Excel (fortschreiben)

- Pro PDF:
    <name>_Inhaltsverzeichnis.pdf
    <name>_Inhaltsverzeichnis.xlsx
- Zusätzlich:
    MASTER_Inhaltsverzeichnis.xlsx (im Ausgabeordner), wird fortgeschrieben.

Templates:
- Muster-Inhaltsverzeichnis-EXCEL.xlsx (wird kopiert und befüllt; Format bleibt identisch)
- Muster-Inhaltsverzeichnis-PDF.pdf (nur als Referenz; PDF wird per ReportLab erzeugt)

Build (Beispiel):
    pyinstaller --onefile --noconsole --name "TOC_Generator" main.py
"""
from __future__ import annotations

import os
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple, Optional

import openpyxl
from pypdf import PdfReader
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.units import mm

import tkinter as tk
from tkinter import filedialog, messagebox

# ----------------------------
# Core: Bookmarks extraction
# ----------------------------

def extract_bookmarks(pdf_path: Path) -> List[Tuple[int, str, Optional[int]]]:
    """Returns list of (level, title, page_number_1_based|None)."""
    reader = PdfReader(str(pdf_path))
    outlines = reader.outline

    items: List[Tuple[int, str, Optional[int]]] = []

    def walk_outline(seq, level: int):
        i = 0
        while i < len(seq):
            it = seq[i]
            if isinstance(it, list):
                walk_outline(it, level + 1)
            else:
                title = getattr(it, "title", str(it))
                page_num = None
                try:
                    page_num = reader.get_destination_page_number(it) + 1
                except Exception:
                    page_num = None
                items.append((level, str(title).strip(), page_num))
                if i + 1 < len(seq) and isinstance(seq[i + 1], list):
                    walk_outline(seq[i + 1], level + 1)
                    i += 1
            i += 1

    if isinstance(outlines, list):
        walk_outline(outlines, 1)
    else:
        # fallback: no list structure
        title = getattr(outlines, "title", str(outlines))
        page_num = None
        try:
            page_num = reader.get_destination_page_number(outlines) + 1
        except Exception:
            page_num = None
        items.append((1, str(title).strip(), page_num))

    return items

# ----------------------------
# PDF output (ReportLab)
# ----------------------------

def draw_toc_pdf(out_pdf: Path, doc_title: str, items: List[Tuple[int, str, Optional[int]]]) -> None:
    width, height = A4
    c = Canvas(str(out_pdf), pagesize=A4)

    left = 25 * mm
    right = 25 * mm
    top = height - 25 * mm
    bottom = 20 * mm

    page_col_x = width - right

    def header() -> float:
        y = top
        c.setFont("Helvetica-Bold", 18)
        c.drawString(left, y, doc_title)
        y -= 20 * mm
        c.setFont("Helvetica", 11)
        c.drawString(left, y, "Inhaltsverzeichnis (aus PDF-Lesezeichen)")
        y -= 10 * mm
        return y

    def str_width(s: str, fontname: str, fontsize: int) -> float:
        return pdfmetrics.stringWidth(s, fontname, fontsize)

    y = header()
    line_h = 7.0 * mm

    for level, title, page in items:
        if y < bottom + line_h:
            c.showPage()
            y = header()

        indent = (level - 1) * 8 * mm
        x_text = left + indent

        # fonts
        if level == 1:
            fontname = "Helvetica-Bold"
            fontsize = 12
        else:
            fontname = "Helvetica"
            fontsize = 12
        c.setFont(fontname, fontsize)

        # right-aligned page number
        page_str = "" if page is None else str(page)
        pw = str_width(page_str, "Helvetica", 12)

        # truncate title if too long
        max_text_width = (page_col_x - 12 * mm) - x_text
        t = title
        while str_width(t, fontname, fontsize) > max_text_width and len(t) > 6:
            t = t[:-2]
        if t != title:
            t = t.rstrip() + "…"

        c.drawString(x_text, y, t)

        # dots leader
        c.setFont("Helvetica", 12)
        dot_char = "."
        dot_w = str_width(dot_char, "Helvetica", 12)

        start = x_text + str_width(t, fontname, fontsize) + 4
        end = page_col_x - pw - 6
        if end > start + dot_w:
            n = int((end - start) / dot_w)
            c.drawString(start, y, dot_char * n)

        # page number
        c.setFont("Helvetica", 12)
        c.drawString(page_col_x - pw, y, page_str)

        y -= line_h

    c.save()

# ----------------------------
# Excel outputs (Template copy)
# ----------------------------

def write_toc_excel(out_xlsx: Path, template_xlsx: Path, doc_title: str, items: List[Tuple[int, str, Optional[int]]]) -> None:
    shutil.copy(template_xlsx, out_xlsx)
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb.active

    # clear existing demo data rows from row 2 down
    for r in range(2, ws.max_row + 1):
        for c in range(1, 4):
            ws.cell(r, c).value = None

    row = 2
    for level, title, page in items:
        ws.cell(row, 1).value = doc_title
        ws.cell(row, 2).value = ("    " * (level - 1)) + title
        ws.cell(row, 3).value = int(page) if page is not None else None
        row += 1

    wb.save(out_xlsx)

def append_master(master_path: Path, template_xlsx: Path, rows: List[Tuple[str, str, Optional[int]]], dedupe: bool = True) -> None:
    if not master_path.exists():
        shutil.copy(template_xlsx, master_path)
        wb = openpyxl.load_workbook(master_path)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            for c in range(1, 4):
                ws.cell(r, c).value = None
        wb.save(master_path)

    wb = openpyxl.load_workbook(master_path)
    ws = wb.active

    existing = set()
    if dedupe:
        for r in range(2, ws.max_row + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            c = ws.cell(r, 3).value
            if a is None and b is None and c is None:
                continue
            existing.add((a, b, c))

    # first empty row
    r = 2
    while any(ws.cell(r, c).value is not None for c in (1, 2, 3)):
        r += 1
        if r > ws.max_row + 20000:
            break

    for rowvals in rows:
        if dedupe and rowvals in existing:
            continue
        ws.cell(r, 1).value = rowvals[0]
        ws.cell(r, 2).value = rowvals[1]
        ws.cell(r, 3).value = rowvals[2]
        r += 1

    wb.save(master_path)

# ----------------------------
# Batch runner
# ----------------------------

@dataclass
class RunConfig:
    input_dir: Path
    output_dir: Path
    template_xlsx: Path
    master_name: str = "MASTER_Inhaltsverzeichnis.xlsx"
    recursive: bool = False
    dedupe_master: bool = True

def iter_pdfs(input_dir: Path, recursive: bool):
    if recursive:
        yield from input_dir.rglob("*.pdf")
    else:
        yield from input_dir.glob("*.pdf")

def run_batch(cfg: RunConfig) -> Tuple[int, int]:
    cfg.output_dir.mkdir(parents=True, exist_ok=True)

    master_path = cfg.output_dir / cfg.master_name
    master_rows: List[Tuple[str, str, Optional[int]]] = []

    ok = 0
    err = 0

    for pdf_path in iter_pdfs(cfg.input_dir, cfg.recursive):
        try:
            items = extract_bookmarks(pdf_path)
            if not items:
                # skip PDFs without outlines
                continue

            doc_title = pdf_path.stem

            out_pdf = cfg.output_dir / f"{doc_title}_Inhaltsverzeichnis.pdf"
            out_xlsx = cfg.output_dir / f"{doc_title}_Inhaltsverzeichnis.xlsx"

            draw_toc_pdf(out_pdf, doc_title, items)
            write_toc_excel(out_xlsx, cfg.template_xlsx, doc_title, items)

            for level, title, page in items:
                master_rows.append((doc_title, ("    " * (level - 1)) + title, int(page) if page else None))

            ok += 1
        except Exception:
            err += 1

    if master_rows:
        append_master(master_path, cfg.template_xlsx, master_rows, dedupe=cfg.dedupe_master)

    return ok, err

# ----------------------------
# GUI
# ----------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Inhaltsverzeichnis aus PDF-Lesezeichen")
        self.geometry("700x300")

        base = Path(__file__).resolve().parent
        self.template_xlsx = base / "Muster-Inhaltsverzeichnis-EXCEL.xlsx"

        self.var_input = tk.StringVar()
        self.var_output = tk.StringVar()
        self.var_recursive = tk.BooleanVar(value=False)
        self.var_dedupe = tk.BooleanVar(value=True)

        frm = tk.Frame(self)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        def row(label, var, cmd):
            r = tk.Frame(frm)
            r.pack(fill="x", pady=4)
            tk.Label(r, text=label, width=20, anchor="w").pack(side="left")
            tk.Entry(r, textvariable=var).pack(side="left", fill="x", expand=True, padx=6)
            tk.Button(r, text="…", command=cmd, width=3).pack(side="left")

        row("Eingabeordner", self.var_input, self.pick_input)
        row("Ausgabeordner", self.var_output, self.pick_output)

        opts = tk.Frame(frm)
        opts.pack(fill="x", pady=8)
        tk.Checkbutton(opts, text="Unterordner einbeziehen (rekursiv)", variable=self.var_recursive).pack(anchor="w")
        tk.Checkbutton(opts, text="Master-Excel: Duplikate verhindern", variable=self.var_dedupe).pack(anchor="w")

        btns = tk.Frame(frm)
        btns.pack(fill="x", pady=10)
        tk.Button(btns, text="Start", command=self.on_start).pack(side="left")
        tk.Button(btns, text="Beenden", command=self.destroy).pack(side="right")

        hint = tk.Label(frm, text="Hinweis: Templates liegen im Programmordner. Master-Excel wird im Ausgabeordner geführt.")
        hint.pack(anchor="w", pady=6)

    def pick_input(self):
        d = filedialog.askdirectory()
        if d:
            self.var_input.set(d)

    def pick_output(self):
        d = filedialog.askdirectory()
        if d:
            self.var_output.set(d)

    def on_start(self):
        inp = self.var_input.get().strip()
        out = self.var_output.get().strip()
        if not inp or not out:
            messagebox.showerror("Fehlt was", "Bitte Eingabe- und Ausgabeordner wählen.")
            return
        if not self.template_xlsx.exists():
            messagebox.showerror("Template fehlt", f"Template nicht gefunden: {self.template_xlsx}")
            return

        cfg = RunConfig(
            input_dir=Path(inp),
            output_dir=Path(out),
            template_xlsx=self.template_xlsx,
            recursive=self.var_recursive.get(),
            dedupe_master=self.var_dedupe.get(),
        )
        ok, err = run_batch(cfg)
        messagebox.showinfo("Fertig", f"Erzeugt: {ok} Inhaltsverzeichnisse\nFehler: {err}\nMaster: {cfg.output_dir / cfg.master_name}")

def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
